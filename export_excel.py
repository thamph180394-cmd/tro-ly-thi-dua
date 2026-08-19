import os
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from database import (
    lay_danh_sach_tuan_thang,
    lay_diem_theo_khoang_thang,
    lay_diem_tu_dau_nam_den_tuan,
    tong_hop_diem_theo_lop_den_tuan,
    tong_hop_diem_theo_lop_khoang_thang,
)

TEN_TRUONG = "TRƯỜNG TH&THCS AN LINH"
TEN_LIEN_DOI = "LIÊN ĐỘI TH&THCS AN LINH"

MAU_XANH_DAM = "0B4EA2"
MAU_XANH_NHAT = "DDEBF7"
MAU_DO = "D71920"
MAU_VANG_NHAT = "FFF2CC"
MAU_XANH_LA_NHAT = "E2F0D9"
MAU_TRANG = "FFFFFF"
MAU_XAM_NHAT = "F7FBFF"
MAU_VIEN = "B7C9E2"

VIEN_MONG = Side(style="thin", color=MAU_VIEN)
BORDER = Border(left=VIEN_MONG, right=VIEN_MONG, top=VIEN_MONG, bottom=VIEN_MONG)


def hien_thi_so(gia_tri):
    try:
        so = float(gia_tri)
        if so.is_integer():
            return int(so)
        return round(so, 2)
    except Exception:
        return gia_tri


def xep_hang_dong_hang(df):
    if df.empty:
        return df

    df = df.sort_values(
        by=["Tổng điểm", "Lớp"],
        ascending=[False, True],
        kind="stable",
    ).reset_index(drop=True)

    df["Hạng"] = (
        df["Tổng điểm"]
        .rank(method="min", ascending=False)
        .astype(int)
    )

    cot_hang = df.pop("Hạng")
    df.insert(0, "Hạng", cot_hang)
    return df


def tao_thu_muc_excel():
    thu_muc = Path("bao_cao_excel")
    thu_muc.mkdir(parents=True, exist_ok=True)
    return thu_muc


def dat_font_mac_dinh(ws):
    for row in ws.iter_rows():
        for cell in row:
            cell.font = Font(name="Times New Roman", size=10)


def dinh_dang_tieu_de(ws, tieu_de, dong_phu, so_cot):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=so_cot)
    c = ws.cell(1, 1, TEN_TRUONG)
    c.font = Font(name="Times New Roman", size=12, bold=True, color=MAU_XANH_DAM)
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=so_cot)
    c = ws.cell(2, 1, TEN_LIEN_DOI)
    c.font = Font(name="Times New Roman", size=11, bold=True, color=MAU_DO)
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=so_cot)
    c = ws.cell(4, 1, tieu_de)
    c.font = Font(name="Times New Roman", size=18, bold=True, color=MAU_XANH_DAM)
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=so_cot)
    c = ws.cell(5, 1, dong_phu)
    c.font = Font(name="Times New Roman", size=12, bold=True)
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[1].height = 21
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[4].height = 28
    ws.row_dimensions[5].height = 20


def dinh_dang_header(ws, dong, so_cot, mau=MAU_XANH_DAM):
    for cot in range(1, so_cot + 1):
        cell = ws.cell(dong, cot)
        cell.fill = PatternFill("solid", fgColor=mau)
        cell.font = Font(name="Times New Roman", size=10, bold=True, color=MAU_TRANG)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[dong].height = 30


def dinh_dang_du_lieu(ws, dong_bat_dau, dong_ket_thuc, so_cot):
    if dong_ket_thuc < dong_bat_dau:
        return

    for dong in range(dong_bat_dau, dong_ket_thuc + 1):
        for cot in range(1, so_cot + 1):
            cell = ws.cell(dong, cot)
            cell.font = Font(name="Times New Roman", size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = BORDER
            if dong % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=MAU_XAM_NHAT)
        ws.row_dimensions[dong].height = 22


def can_chinh_cot(ws, widths):
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width


def tao_dataframe_xep_hang(du_lieu):
    cot = [
        "Lớp", "Khối", "Số tuần", "Học tập", "Kỷ luật", "Vệ sinh",
        "Điểm cộng", "Điểm trừ", "Tổng điểm", "Điểm TB",
    ]
    df = pd.DataFrame(du_lieu, columns=cot)
    if df.empty:
        return df

    for cot_so in [
        "Số tuần", "Học tập", "Kỷ luật", "Vệ sinh", "Điểm cộng",
        "Điểm trừ", "Tổng điểm", "Điểm TB",
    ]:
        df[cot_so] = pd.to_numeric(df[cot_so], errors="coerce").fillna(0)

    return xep_hang_dong_hang(df)


def tao_sheet_xep_hang(wb, df, tieu_de, dong_phu):
    ws = wb.active
    ws.title = "Xếp hạng"

    headers = [
        "Hạng", "Lớp", "Khối", "Số tuần", "Học tập", "Kỷ luật",
        "Vệ sinh", "Điểm cộng", "Điểm trừ", "Tổng điểm", "Điểm TB",
    ]

    dinh_dang_tieu_de(ws, tieu_de, dong_phu, len(headers))

    tong_lop = len(df)
    diem_tb = round(df["Điểm TB"].mean(), 2) if not df.empty else 0
    if not df.empty:
        hang_1 = df[df["Hạng"] == 1]
        lop_dan_dau = ", ".join(hang_1["Lớp"].astype(str).tolist())
        diem_dan_dau = hang_1.iloc[0]["Tổng điểm"]
    else:
        lop_dan_dau = ""
        diem_dan_dau = 0

    tong_quan = [
        ("A7", "TỔNG SỐ LỚP", "B7", tong_lop),
        ("D7", "ĐIỂM TB/TUẦN", "E7", diem_tb),
        ("G7", "LỚP DẪN ĐẦU", "H7", lop_dan_dau),
        ("J7", "TỔNG ĐIỂM", "K7", hien_thi_so(diem_dan_dau)),
    ]
    for label_cell, label, value_cell, value in tong_quan:
        ws[label_cell] = label
        ws[label_cell].fill = PatternFill("solid", fgColor=MAU_XANH_NHAT)
        ws[label_cell].font = Font(name="Times New Roman", size=10, bold=True, color=MAU_XANH_DAM)
        ws[label_cell].alignment = Alignment(horizontal="center", vertical="center")
        ws[label_cell].border = BORDER

        ws[value_cell] = value
        ws[value_cell].font = Font(name="Times New Roman", size=11, bold=True, color=MAU_DO)
        ws[value_cell].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws[value_cell].border = BORDER

    dong_header = 10
    for i, header in enumerate(headers, start=1):
        ws.cell(dong_header, i, header)
    dinh_dang_header(ws, dong_header, len(headers))

    dong = dong_header + 1
    for _, row in df.iterrows():
        values = [
            row["Hạng"], row["Lớp"], row["Khối"], row["Số tuần"],
            row["Học tập"], row["Kỷ luật"], row["Vệ sinh"], row["Điểm cộng"],
            row["Điểm trừ"], row["Tổng điểm"], round(row["Điểm TB"], 2),
        ]
        for cot, value in enumerate(values, start=1):
            ws.cell(dong, cot, hien_thi_so(value))
        dong += 1

    dinh_dang_du_lieu(ws, dong_header + 1, dong - 1, len(headers))

    for r in range(dong_header + 1, dong):
        try:
            hang = int(ws.cell(r, 1).value)
        except Exception:
            continue

        if hang <= 5:
            for c in range(1, len(headers) + 1):
                ws.cell(r, c).fill = PatternFill("solid", fgColor=MAU_VANG_NHAT)

        if hang == 1:
            ws.cell(r, 1).font = Font(name="Times New Roman", size=10, bold=True, color="C47A00")
            ws.cell(r, 10).font = Font(name="Times New Roman", size=10, bold=True, color=MAU_DO)

    can_chinh_cot(ws, [8, 12, 12, 10, 12, 12, 12, 12, 12, 14, 12])
    ws.freeze_panes = "A11"
    ws.sheet_view.showGridLines = False
    ws.auto_filter.ref = f"A10:K{max(10, dong - 1)}"

    if len(df) > 0:
        chart = BarChart()
        chart.type = "bar"
        chart.style = 10
        chart.title = "TỔNG ĐIỂM THI ĐUA CÁC LỚP"
        chart.y_axis.title = "Lớp"
        chart.x_axis.title = "Tổng điểm"

        data = Reference(ws, min_col=10, min_row=10, max_row=dong - 1)
        cats = Reference(ws, min_col=2, min_row=11, max_row=dong - 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 8
        chart.width = 16
        ws.add_chart(chart, "M10")


def tao_du_lieu_thang_tu_chi_tiet(chi_tiet, khoi_map=None):
    cot = [
        "Tuần", "Tháng", "Lớp", "Học tập", "Kỷ luật", "Vệ sinh",
        "Điểm cộng", "Điểm trừ", "Tổng điểm", "Ghi chú",
    ]
    df = pd.DataFrame(chi_tiet, columns=cot)
    if df.empty:
        return []

    df["Tổng điểm"] = pd.to_numeric(df["Tổng điểm"], errors="coerce").fillna(0)
    df = df.dropna(subset=["Tháng"])
    if df.empty:
        return []

    df["Tháng"] = df["Tháng"].astype(int)
    tong = (
        df.groupby(["Lớp", "Tháng"], as_index=False)["Tổng điểm"]
        .sum()
    )

    khoi_map = khoi_map or {}
    ket_qua = []
    for _, row in tong.iterrows():
        lop = str(row["Lớp"])
        ket_qua.append((lop, khoi_map.get(lop, ""), int(row["Tháng"]), float(row["Tổng điểm"])))
    return ket_qua


def tao_sheet_theo_thang(wb, nam_hoc, du_lieu, danh_sach_thang, tieu_de):
    ws = wb.create_sheet("Tổng hợp theo tháng")
    df = pd.DataFrame(du_lieu, columns=["Lớp", "Khối", "Tháng", "Tổng điểm tháng"])

    headers = ["Hạng", "Lớp", "Khối"] + [f"Tháng {t}" for t in danh_sach_thang] + ["Tổng điểm"]
    dinh_dang_tieu_de(ws, tieu_de, f"Năm học {nam_hoc}", len(headers))

    if df.empty:
        ws["A8"] = "Chưa có dữ liệu theo tháng. Hãy kiểm tra phần gán Tuần - Tháng."
        ws["A8"].font = Font(name="Times New Roman", size=11, italic=True, color=MAU_DO)
        can_chinh_cot(ws, [18] * max(1, len(headers)))
        ws.sheet_view.showGridLines = False
        return

    pivot = df.pivot_table(
        index=["Lớp", "Khối"],
        columns="Tháng",
        values="Tổng điểm tháng",
        aggfunc="sum",
        fill_value=0,
    ).reset_index()

    for thang in danh_sach_thang:
        if thang not in pivot.columns:
            pivot[thang] = 0

    pivot["Tổng điểm"] = pivot[danh_sach_thang].sum(axis=1)
    pivot = pivot.sort_values(by=["Tổng điểm", "Lớp"], ascending=[False, True], kind="stable").reset_index(drop=True)
    pivot["Hạng"] = pivot["Tổng điểm"].rank(method="min", ascending=False).astype(int)

    dong_header = 8
    for i, header in enumerate(headers, start=1):
        ws.cell(dong_header, i, header)
    dinh_dang_header(ws, dong_header, len(headers))

    dong = dong_header + 1
    for _, row in pivot.iterrows():
        values = [row["Hạng"], row["Lớp"], row["Khối"]]
        values += [row.get(thang, 0) for thang in danh_sach_thang]
        values += [row["Tổng điểm"]]

        for c, value in enumerate(values, start=1):
            ws.cell(dong, c, hien_thi_so(value))
        dong += 1

    dinh_dang_du_lieu(ws, dong_header + 1, dong - 1, len(headers))

    for r in range(dong_header + 1, dong):
        try:
            hang = int(ws.cell(r, 1).value)
        except Exception:
            continue
        if hang <= 5:
            for c in range(1, len(headers) + 1):
                ws.cell(r, c).fill = PatternFill("solid", fgColor=MAU_VANG_NHAT)

    widths = [8, 12, 12] + [14 for _ in danh_sach_thang] + [15]
    can_chinh_cot(ws, widths)
    ws.freeze_panes = "A9"
    ws.auto_filter.ref = f"A8:{get_column_letter(len(headers))}{max(8, dong - 1)}"
    ws.sheet_view.showGridLines = False


def tao_sheet_chi_tiet(wb, nam_hoc, du_lieu, tieu_de, dong_phu):
    ws = wb.create_sheet("Chi tiết từng tuần")
    headers = [
        "Tuần", "Tháng", "Lớp", "Học tập", "Kỷ luật", "Vệ sinh",
        "Điểm cộng", "Điểm trừ", "Tổng điểm", "Ghi chú",
    ]

    dinh_dang_tieu_de(ws, tieu_de, dong_phu, len(headers))

    dong_header = 8
    for i, header in enumerate(headers, start=1):
        ws.cell(dong_header, i, header)
    dinh_dang_header(ws, dong_header, len(headers))

    dong = dong_header + 1
    for item in du_lieu:
        values = list(item)
        for c, value in enumerate(values, start=1):
            ws.cell(dong, c, hien_thi_so(value))
        dong += 1

    dinh_dang_du_lieu(ws, dong_header + 1, dong - 1, len(headers))
    can_chinh_cot(ws, [9, 9, 12, 12, 12, 12, 12, 12, 14, 34])
    ws.freeze_panes = "A9"
    ws.auto_filter.ref = f"A8:J{max(8, dong - 1)}"
    ws.sheet_view.showGridLines = False


def tao_sheet_lich_tuan(wb, nam_hoc):
    ws = wb.create_sheet("Lịch tuần - tháng")
    du_lieu = lay_danh_sach_tuan_thang(nam_hoc)

    dinh_dang_tieu_de(ws, "LỊCH TUẦN - THÁNG", f"Năm học {nam_hoc}", 2)
    ws["A8"] = "Tuần"
    ws["B8"] = "Tháng"
    dinh_dang_header(ws, 8, 2)

    dong = 9
    for tuan, thang in du_lieu:
        ws.cell(dong, 1, int(tuan))
        ws.cell(dong, 2, f"Tháng {int(thang)}")
        dong += 1

    dinh_dang_du_lieu(ws, 9, dong - 1, 2)
    can_chinh_cot(ws, [15, 20])
    ws.sheet_view.showGridLines = False


def tao_excel_so_ket(nam_hoc):
    """
    Sơ kết thi đua từ Tháng 9 đến hết Tháng 12.
    """
    du_lieu_xep_hang = tong_hop_diem_theo_lop_khoang_thang(nam_hoc, 9, 12)
    if not du_lieu_xep_hang:
        raise ValueError(
            "Chưa có dữ liệu Sơ kết từ Tháng 9 đến Tháng 12. "
            "Hãy kiểm tra các tuần đã được gán đúng tháng chưa."
        )

    df = tao_dataframe_xep_hang(du_lieu_xep_hang)
    chi_tiet = lay_diem_theo_khoang_thang(nam_hoc, 9, 12)
    khoi_map = dict(zip(df["Lớp"].astype(str), df["Khối"].astype(str)))
    du_lieu_thang = tao_du_lieu_thang_tu_chi_tiet(chi_tiet, khoi_map)

    wb = Workbook()
    tao_sheet_xep_hang(
        wb,
        df,
        "SƠ KẾT THI ĐUA TOÀN TRƯỜNG",
        f"Năm học {nam_hoc} - Từ Tháng 9 đến hết Tháng 12",
    )
    tao_sheet_theo_thang(
        wb,
        nam_hoc,
        du_lieu_thang,
        [9, 10, 11, 12],
        "TỔNG HỢP ĐIỂM THI ĐUA THEO THÁNG",
    )
    tao_sheet_chi_tiet(
        wb,
        nam_hoc,
        chi_tiet,
        "CHI TIẾT ĐIỂM THI ĐUA TỪ THÁNG 9 ĐẾN THÁNG 12",
        f"Năm học {nam_hoc}",
    )
    tao_sheet_lich_tuan(wb, nam_hoc)

    thu_muc = tao_thu_muc_excel()
    ten_nam_hoc = nam_hoc.replace(" ", "").replace("-", "_")
    duong_dan = (thu_muc / f"so_ket_thi_dua_{ten_nam_hoc}.xlsx").resolve()
    wb.save(duong_dan)
    return str(duong_dan)


def tao_excel_tong_ket(nam_hoc, tuan_hien_tai):
    """
    Tổng kết thi đua từ đầu năm học đến Tuần hiện tại.
    Khi đang ở tháng 5, chỉ lấy đến đúng tuần người dùng đang chọn.
    """
    tuan_hien_tai = int(tuan_hien_tai)

    du_lieu_xep_hang = tong_hop_diem_theo_lop_den_tuan(nam_hoc, tuan_hien_tai)
    if not du_lieu_xep_hang:
        raise ValueError("Chưa có dữ liệu Tổng kết thi đua đến tuần hiện tại.")

    df = tao_dataframe_xep_hang(du_lieu_xep_hang)
    chi_tiet = lay_diem_tu_dau_nam_den_tuan(nam_hoc, tuan_hien_tai)

    khoi_map = dict(zip(df["Lớp"].astype(str), df["Khối"].astype(str)))
    du_lieu_thang = tao_du_lieu_thang_tu_chi_tiet(chi_tiet, khoi_map)

    cac_thang_da_co = []
    for _, thang in lay_danh_sach_tuan_thang(nam_hoc):
        thang = int(thang)
        if thang not in cac_thang_da_co:
            cac_thang_da_co.append(thang)

    thu_tu_nam_hoc = [9, 10, 11, 12, 1, 2, 3, 4, 5]
    danh_sach_thang = [t for t in thu_tu_nam_hoc if t in cac_thang_da_co]
    if not danh_sach_thang:
        danh_sach_thang = thu_tu_nam_hoc

    wb = Workbook()
    tao_sheet_xep_hang(
        wb,
        df,
        "TỔNG KẾT THI ĐUA TOÀN TRƯỜNG",
        f"Năm học {nam_hoc} - Từ đầu năm đến Tuần {tuan_hien_tai}",
    )
    tao_sheet_theo_thang(
        wb,
        nam_hoc,
        du_lieu_thang,
        danh_sach_thang,
        "TỔNG HỢP ĐIỂM THI ĐUA THEO THÁNG",
    )
    tao_sheet_chi_tiet(
        wb,
        nam_hoc,
        chi_tiet,
        "CHI TIẾT ĐIỂM THI ĐUA TỪ ĐẦU NĂM",
        f"Năm học {nam_hoc} - Đến Tuần {tuan_hien_tai}",
    )
    tao_sheet_lich_tuan(wb, nam_hoc)

    thu_muc = tao_thu_muc_excel()
    ten_nam_hoc = nam_hoc.replace(" ", "").replace("-", "_")
    duong_dan = (
        thu_muc
        / f"tong_ket_thi_dua_{ten_nam_hoc}_den_tuan_{tuan_hien_tai}.xlsx"
    ).resolve()
    wb.save(duong_dan)
    return str(duong_dan)


if __name__ == "__main__":
    print("==============================================")
    print("FILE EXPORT_EXCEL.PY ĐÃ SẴN SÀNG")
    print("==============================================")
    print()
    print("Sơ kết thi đua:")
    print("tao_excel_so_ket(nam_hoc)")
    print()
    print("Tổng kết thi đua:")
    print("tao_excel_tong_ket(nam_hoc, tuan_hien_tai)")
    print()
    print("==============================================")
